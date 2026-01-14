from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl.styles import PatternFill


# ----------------------------
# Filtering rules (license.csv)
# ----------------------------
TIERS_ALLOWED = {"enterprise", "enterprise ai", "inspect pro", "starter+"}
PRODUCT_ALLOWED = "inspect"

BU_EXCLUDE = "proceq hq"
BU_MISSING_MARKERS = {"", "#n/a", "n/a", "na", "none", "null"}


# ----------------------------
# Excel formatting
# ----------------------------
ORANGE_FILL = PatternFill(start_color="FFFFA500", end_color="FFFFA500", fill_type="solid")  # Orange
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")  # Yellow


def project_root_from_this_file() -> Path:
    """
    Returns the project root, assuming this file lives at:
      <root>/src/project_renewal/consolidate.py
    """
    return Path(__file__).resolve().parents[2]


def pick_latest_csv_by_prefix(data_dir: Path, prefix: str) -> Path:
    """
    Find the latest CSV file in data_dir whose filename starts with `prefix` (case-insensitive).
    Chooses by modification time (mtime). Raises if none found.

    Example matches:
      contracts_saas2026-01-14_01_25.csv  (prefix: "contracts")
      licenses_saas2026-01-14_01_26.csv   (prefix: "licenses")
    """
    prefix_lower = prefix.lower()

    matches = [
        p for p in data_dir.glob("*.csv")
        if p.name.lower().startswith(prefix_lower)
    ]

    if not matches:
        raise FileNotFoundError(
            f"No CSV found in {data_dir} starting with '{prefix}'. "
            f"Available CSVs: {[p.name for p in data_dir.glob('*.csv')]}"
        )

    # pick newest by modified time
    matches.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return matches[0]


def read_standard_csv(path: Path) -> pd.DataFrame:
    df = pd.read_csv(
        path,
        encoding="utf-8-sig",
        dtype=str,
        keep_default_na=False,
    )
    df.columns = [c.strip() for c in df.columns]
    return df


def read_license_csv(path: Path) -> pd.DataFrame:
    """
    license.csv in your earlier sample had an extra first line like:
        Contract,,,,,License,,,,,,
    so we try skiprows=1 first.
    If that doesn't yield expected columns, fall back to normal read.
    """
    # First attempt: skip decorative first row
    df1 = pd.read_csv(
        path,
        encoding="utf-8-sig",
        skiprows=1,
        dtype=str,
        keep_default_na=False,
    )
    df1.columns = [c.strip() for c in df1.columns]

    # If this looks valid, keep it; otherwise retry without skipping.
    expected = {"Contract ID", "Product", "Tier"}
    if expected.issubset(set(df1.columns)):
        return df1

    # Fallback: read as a normal CSV
    df2 = pd.read_csv(
        path,
        encoding="utf-8-sig",
        dtype=str,
        keep_default_na=False,
    )
    df2.columns = [c.strip() for c in df2.columns]
    return df2


def filter_license_rows(license_df: pd.DataFrame) -> pd.DataFrame:
    """
    Filter license.csv:
      - Product == Inspect (case-insensitive)
      - Tier in {Enterprise, Enterprise AI, Inspect Pro, Starter+} (case-insensitive)
      - BU NOT in {"Proceq HQ", "", "#N/A", "N/A", "NA", "None", "Null"} (case-insensitive)
    """
    required_cols = {"Contract ID", "Product", "Tier", "BU"}
    missing = required_cols - set(license_df.columns)
    if missing:
        raise ValueError(f"license file is missing columns: {sorted(missing)}")

    product_norm = license_df["Product"].astype(str).str.strip().str.lower()
    tier_norm = license_df["Tier"].astype(str).str.strip().str.lower()
    bu_norm = license_df["BU"].astype(str).str.strip().str.lower()

    mask = (
        product_norm.eq(PRODUCT_ALLOWED)
        & tier_norm.isin(TIERS_ALLOWED)
        & ~bu_norm.eq(BU_EXCLUDE)
        & ~bu_norm.isin(BU_MISSING_MARKERS)
    )
    return license_df.loc[mask].copy()


def build_contract_lookup(contract_df: pd.DataFrame) -> Dict[str, Dict[str, str]]:
    """
    Key:
      contract_lookup[contract_id] where contract_id = contract_df["ID"]

    Values added to each license item:
      Country Sold To, User Type, Remarks, BP, First Expiration Date, License Count, Language
    """
    required_cols = {
        "ID",
        "Country Sold To",
        "User Type",
        "Remarks",
        "BP",
        "First Expiration Date",
        "License Count",
        "Language",
    }
    missing = required_cols - set(contract_df.columns)
    if missing:
        raise ValueError(f"contract file is missing columns: {sorted(missing)}")

    fields = [
        "Country Sold To",
        "User Type",
        "Remarks",
        "BP",
        "First Expiration Date",
        "License Count",
        "Language",
    ]

    lookup: Dict[str, Dict[str, str]] = {}
    for _, row in contract_df.iterrows():
        cid = str(row["ID"]).strip()
        if not cid or cid in lookup:
            continue
        lookup[cid] = {f: str(row.get(f, "")).strip() for f in fields}
    return lookup


def build_finance_lookup(finance_df: pd.DataFrame) -> Dict[str, str]:
    """
    Key:
      finance_lookup[contract_id] where contract_id = finance_df["Contract ID"]

    Value:
      Customer name (default handled by caller as #NA if not found)
    """
    required_cols = {"Contract ID", "Customer name"}
    missing = required_cols - set(finance_df.columns)
    if missing:
        raise ValueError(f"finance file is missing columns: {sorted(missing)}")

    lookup: Dict[str, str] = {}
    for _, row in finance_df.iterrows():
        cid = str(row["Contract ID"]).strip()
        cname = str(row.get("Customer name", "")).strip()
        if not cid:
            continue

        # keep first non-empty name
        if cid in lookup and lookup[cid] != "#NA":
            continue

        lookup[cid] = cname if cname else lookup.get(cid, "#NA")

    return lookup


def autofit_excel_columns(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    """
    Auto-fit column widths in the given Excel sheet based on cell content length.
    Works with openpyxl engine.
    """
    ws = writer.sheets[sheet_name]
    for idx, col in enumerate(df.columns, start=1):
        series = df[col].astype(str).fillna("")
        max_len = max([len(str(col))] + series.map(len).tolist())
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(max_len + 2, 60)


def highlight_expiring_rows(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame, date_col: str = "Expiration") -> None:
    """
    Highlight entire rows based on Expiration date:
      - 0–30 days from today: orange
      - 31–60 days from today: yellow
    """
    if date_col not in df.columns:
        return

    ws = writer.sheets[sheet_name]
    exp = pd.to_datetime(df[date_col], errors="coerce")
    today = pd.Timestamp.today().normalize()

    for excel_row_idx, exp_ts in enumerate(exp, start=2):  # row 1 is header
        if pd.isna(exp_ts):
            continue
        days_to_expiry = (exp_ts.normalize() - today).days

        if 0 <= days_to_expiry <= 30:
            fill = ORANGE_FILL
        elif 31 <= days_to_expiry <= 60:
            fill = YELLOW_FILL
        else:
            continue

        for col_idx in range(1, df.shape[1] + 1):
            ws.cell(row=excel_row_idx, column=col_idx).fill = fill


def consolidate(data_dir: Path) -> Path:
    # ---- Find latest input files by prefix ----
    contract_path = pick_latest_csv_by_prefix(data_dir, "contracts")
    license_path = pick_latest_csv_by_prefix(data_dir, "licenses")

    # finance: keep backward compatibility with "finance.csv"
    # If you also use dated files, this will still work.
    finance_path: Optional[Path] = None
    for prefix in ("finance", "finances"):
        try:
            finance_path = pick_latest_csv_by_prefix(data_dir, prefix)
            break
        except FileNotFoundError:
            continue
    if finance_path is None:
        # fallback to exact name if present
        fp = data_dir / "finance.csv"
        if fp.exists():
            finance_path = fp
        else:
            raise FileNotFoundError(
                f"No finance file found. Expected a CSV starting with 'finance'/'finances' "
                f"or a file named 'finance.csv' in {data_dir}."
            )

    output_path = data_dir / "consolidated.xlsx"

    # ---- Read inputs ----
    lic_df = read_license_csv(license_path)
    con_df = read_standard_csv(contract_path)
    fin_df = read_standard_csv(finance_path)

    # ---- Filter licenses ----
    lic_filtered = filter_license_rows(lic_df)

    # ---- Create data structure (list of dicts) ----
    items: List[Dict[str, Any]] = lic_filtered.to_dict(orient="records")

    # ---- Build lookups ----
    contract_lookup = build_contract_lookup(con_df)
    finance_lookup = build_finance_lookup(fin_df)

    contract_fields = [
        "Country Sold To",
        "User Type",
        "Remarks",
        "BP",
        "First Expiration Date",
        "License Count",
        "Language",
    ]

    missing_contract = 0
    missing_finance = 0

    # ---- Enrich each license record ----
    for item in items:
        contract_id = str(item.get("Contract ID", "")).strip()

        # Contract fields: license["Contract ID"] -> contract["ID"]
        cinfo = contract_lookup.get(contract_id)
        if cinfo is None:
            missing_contract += 1
            for f in contract_fields:
                item.setdefault(f, "")
        else:
            item.update(cinfo)

        # Finance: license["Contract ID"] -> finance["Contract ID"]
        customer_name = finance_lookup.get(contract_id, "#NA")
        if customer_name == "#NA":
            missing_finance += 1
        item["Customer name"] = customer_name

    # ---- Output DF ----
    out_df = pd.DataFrame(items)

    # Sort by Expiration (invalid/missing last)
    if "Expiration" in out_df.columns:
        _exp_sort = pd.to_datetime(out_df["Expiration"], errors="coerce")
        out_df = (
            out_df.assign(_exp_sort=_exp_sort)
            .sort_values("_exp_sort", na_position="last")
            .drop(columns=["_exp_sort"])
        )

    # Ensure "Customer name" is column 2
    if "Customer name" not in out_df.columns:
        out_df["Customer name"] = "#NA"

    cols = list(out_df.columns)
    first_col = "Contract ID" if "Contract ID" in cols else next((c for c in cols if c != "Customer name"), cols[0])
    new_cols = [first_col, "Customer name"] + [c for c in cols if c not in {first_col, "Customer name"}]
    out_df = out_df[new_cols]

    # Write Excel with highlights
    data_dir.mkdir(parents=True, exist_ok=True)

    sheet_name = "consolidated"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        out_df.to_excel(writer, index=False, sheet_name=sheet_name)
        highlight_expiring_rows(writer, sheet_name, out_df, date_col="Expiration")
        autofit_excel_columns(writer, sheet_name, out_df)

    print("Inputs used:")
    print(f"  contracts: {contract_path.name}")
    print(f"  licenses : {license_path.name}")
    print(f"  finance  : {finance_path.name}")
    print("")
    print(f"Filtered licenses: {len(lic_filtered):,} rows")
    print(f"Missing contract matches (ID not found in contracts file): {missing_contract:,}")
    print(f"Missing finance matches (Customer name defaulted to #NA): {missing_finance:,}")
    print(f"Output written to: {output_path}")

    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Consolidate contracts/licenses/finance CSVs into data/consolidated.xlsx"
    )
    parser.add_argument(
        "--data-dir",
        type=str,
        default=None,
        help="Path to the data folder (default: <project_root>/data).",
    )
    args = parser.parse_args()

    root = project_root_from_this_file()
    data_dir = Path(args.data_dir) if args.data_dir else (root / "data")

    consolidate(data_dir)


if __name__ == "__main__":
    main()
